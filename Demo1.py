import psycopg2
import pandas as pd
from psycopg2.extras import RealDictCursor
from openpyxl import load_workbook
import numpy as np

hostname ='localhost'
database ='Akara'
username ='postgres'
pwd ='zcon@123'
port_id =5432
conn = None
cur = None
try:
        conn = psycopg2.connect(
               host = hostname,
                dbname = database,
                user = username,
                password = pwd,
                port = port_id)
        cur = conn.cursor()
        cur.execute("select company,brand,ndc11,initial_date,termination_date,product_strength,product_volume,units_per_package,market_basket,dosage_form,is_line_extension from launch_product") 
        DATA = (cur.fetchall())
        df = pd.DataFrame(DATA)
        writer = pd.ExcelWriter('Data.xlsx',engine = 'xlsxwriter')
        a=(len(df))
        df.to_excel("Data.xlsx",sheet_name='Launch_Product',na_rep='-', float_format=None, columns=None, header=False, index=True, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, inf_rep='inf', freeze_panes=(1, 1), storage_options=None)
        wb = load_workbook('Data.xlsx')
        ws = wb['Launch_Product']
        for i in range(1,a+1):
              B=str("A")+str(i)
              ws[B] = ""
              wb.save('Data.xlsx')
        print("sheet created Launch_Product ")
        cur=conn.cursor()
        
        
        cur.execute("(select to_char(date, 'MonYY'),topline_units,actual_topline_units,adjusted_topline_units,patients,ta_market_basket_units,ta_market_basket_units_change from launch_national_patient where product_id = 1089788)" )
        DATA1 = (cur.fetchall())
        #print(tuple(DATA))
        DATA2=tuple(zip(*DATA1))
        df1= pd.DataFrame(DATA2)
        wb1 = load_workbook('Data.xlsx')
        wb1.create_sheet('National_Patient')
        writer=pd.ExcelWriter("Data.xlsx",engine="openpyxl",mode="a")
        df1.to_excel(writer,sheet_name='National_Patient', na_rep='-', float_format=None, columns=None, header=False, index=True, index_label=None, startrow=0, startcol=0, engine = 'xlsxwriter',merge_cells=True, inf_rep='inf', freeze_panes=None, storage_options=None)
        writer.close()
        wb2 = load_workbook('Data.xlsx')
        ws2= wb2['National_Patient']
        ws2['A1'] = ""
        ws2['A2'] = "Topline Units"
        ws2['A3'] = "Actual Topline Units"
        ws2['A4'] = "Adjusted Units"
        ws2['A5'] = "Patients"
        ws2['A6'] = "TA Market Basket Units"
        ws2['A7'] = "TA Market Basket Units % change"
        wb2.save('Data.xlsx')
        print("sheet created National_Patient ")
        cur.close()
        conn.close()

except Exception as error:
    print(error)